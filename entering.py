import os
import django
import csv

# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'trial.settings')
django.setup()
from exam.models import phy_easy


#q = phy_easy.objects.get(pk=2)
#q.question="alfred"
#q.save()
print(phy_easy.objects.all())

from openpyxl import load_workbook
excel_file = "Book1.xlsx"
wb = load_workbook(excel_file)
ws = wb.active
for row in ws.iter_rows(min_row=1, values_only=True):
    classs,category,chapter, opt1, opt2, opt3, opt4,question,questionimg, answer, danswer, danswerimg = row
    #print(question) 
    #iltered_blogs = phy_easy.objects.filter(category="physics")
    #filtered_blogs.delete()    
    #phy_easy.objects.create(question=question, question_img=questionimg, classs=classs, category=category, chapter=chapter,answer=answer, danswer= danswer, danswer_img=danswerimg, option1=opt1, option2=opt2, option3=opt3, option4= opt4)
    #print(phy_easy.objects.all())

#blog = Blog(name="My Blog", tagline="Blog Tagline")
#blog.save()

#all_blogs = Blog.objects.all()
#single_blog = Blog.objects.get(id=1)


print(phy_easy.objects.all())

#blog = Blog.objects.get(id=1)
#blog.name = "Updated Blog Name"
#blog.save()

my_instance = phy_easy.objects.get(id=1)
print(my_instance)
#blog.delete()

#from django.db.models import Q
#complex_query = Blog.objects.filter(Q(name__startswith="D") | Q(tagline__icontains="django"))
