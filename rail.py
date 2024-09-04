import pandas as pd
from docx2python import docx2python

from pypdf import PdfReader
from openpyxl import load_workbook


def extract_pages(file_path):
    reader = PdfReader(file_path)
    pages = []
    pages.append(namm)
    for i in range(len(reader.pages)):
        page = reader.pages[i]
        text = page.extract_text()
        #pages.append({"page_no": i + 1, "page_text": text})
        #pages.append({"A": text})# change column while entering next chap
        pages.append(text)# change column while entering next chap
    return pages

# Load the existing workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# Example usage
namm = 'Kinetic Theory of Gases'
namm = 'Electrostatics'
namm = 'Laws of Motion'
namm = 'Gravitation'
namm = 'Work,Energy and Power'
namm = 'Center of Mass and Momentum Conservation'
namm = 'System of Particles'
namm = 'Oscillation'
file_path = 'E:/'+namm+'.pdf'
pages = extract_pages(file_path)
#for page in pages:
   #ws.append(page)
   #print(page)

# Write the list to the first column
for row, value in enumerate(pages, start=1):
    ws.cell(row=row, column=8, value=value)# change column value

# Save the workbook
wb.save("sample.xlsx")

conte=[]
print("done")
#df=pd.read_excel("sample.xlsx")
#conte= df['heat'].tolist()
 
